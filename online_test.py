import streamlit as st
import calendar
from datetime import date, datetime, time, timedelta
import holidays
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
import io

def heures_vers_texte(nb_heures):
    heures = int(nb_heures)
    minutes = int(round((nb_heures - heures) * 60))
    return f"{heures:02d}:{minutes:02d}"

def calculer_heures(ws, ligne_debut, ligne_fin, col_heure_debut, col_heure_fin, col_resultat):
    fmt = "%H:%M"
    for row in range(ligne_debut, ligne_fin + 1):
        debut = ws.cell(row=row, column=col_heure_debut).value
        fin = ws.cell(row=row, column=col_heure_fin).value

        if isinstance(debut, str) and isinstance(fin, str):
            try:
                h_debut = datetime.strptime(debut, fmt)
                h_fin = datetime.strptime(fin, fmt)
                diff = (h_fin - h_debut).seconds / 3600
                ws.cell(row=row, column=col_resultat, value=heures_vers_texte(diff))
            except ValueError:
                ws.cell(row=row, column=col_resultat, value="")
        else:
            ws.cell(row=row, column=col_resultat, value="")

# Pour regrouper plusieurs jours de congés, d'absences ou d'arret maladies qui se suivent
def regrouper_plages(vacances):
    if not vacances:
        return []

    dates = sorted(vacances.keys())
    plages = []

    debut = dates[0]
    fin = dates[0]

    for d in dates[1:]:
        if d == fin + timedelta(days=1):
            fin = d
        else:
            plages.append((debut, fin))
            debut = d
            fin = d

    plages.append((debut, fin))
    return plages

# Fais la somme des heures travaillées (donc 3e colonne) pour chaque semaine ou pour le mois entier
def somme(ws, lignes, col_source, col_resultat, ligne_total, somme):
    total_minutes = 0

    if somme == "semaine":
        for row in lignes:
            val = ws.cell(row=row, column=col_source).value
            if isinstance(val, str):
                try:
                    heures, minutes = map(int, val.split(":"))
                    total_minutes += heures * 60 + minutes
                except:
                    pass

    elif somme == "total":
        for col in col_source:
            val = ws.cell(row=lignes, column=col).value
            if isinstance(val, str):
                try:
                    heures, minutes = map(int, val.split(":"))
                    total_minutes += heures * 60 + minutes
                except:
                    pass

    total_heures = total_minutes // 60
    reste_minutes = total_minutes % 60
    ws.cell(row=ligne_total, column=col_resultat, value=f"{total_heures:02d}:{reste_minutes:02d}")

def remplir_calendrier(ws, mois, annee, vacances, absences, arret, nom, responsable, ddc, fdc, vacances_total, absences_total, arret_total):
    mois_string = ["JANVIER", "FEVRIER", "MARS", "AVRIL", "MAI", "JUIN", "JUILLET", "AOUT", "SEPTEMBRE", "OCTOBRE", "NOVEMBRE", "DECEMBRE"]

    # Remplissage des infos de base

    ws.cell(row=2, column=27, value=nom)
    ws.cell(row=6, column=27, value=responsable)
    ws.cell(row=2, column=13, value=mois_string[mois-1])
    ws.cell(row=5, column=13, value=annee)
    ws.cell(row=30, column=9, value=vacances_total)
    ws.cell(row=27, column=9, value=absences_total)
    ws.cell(row=27, column=19, value=arret_total)
    ws.cell(row=30, column=28, value=ddc.strftime("%x"))
    if (isinstance(fdc, date)):
        ws.cell(row=35, column=28, value=fdc.strftime("%x"))
    else:
        ws.cell(row=35, column=28, value=fdc)

    # Remplissage des groupes de CP, ABS et AM dans la cartouche du bas

    plages_vacances = regrouper_plages(vacances)
    nb_plages_vac = 0

    for d1, d2 in plages_vacances:
        ws.cell(row=30+nb_plages_vac, column=2, value="du")
        ws.cell(row=30+nb_plages_vac, column=3, value=f"{d1.strftime('%d/%m')}")
        ws.cell(row=30+nb_plages_vac, column=4, value="au")
        ws.cell(row=30+nb_plages_vac, column=5, value=f"{d2.strftime('%d/%m')}")
        nb_plages_vac += 1

    plages_absences = regrouper_plages(absences)
    nb_plages_abs = 0

    for d1, d2 in plages_absences:
        ws.cell(row=27+nb_plages_abs, column=2, value="du")
        ws.cell(row=27+nb_plages_abs, column=3, value=f"{d1.strftime('%d/%m')}")
        ws.cell(row=27+nb_plages_abs, column=4, value="au")
        ws.cell(row=27+nb_plages_abs, column=5, value=f"{d2.strftime('%d/%m')}")
        nb_plages_abs += 1
    
    plages_arret = regrouper_plages(arret)
    nb_plages_arret = 0

    for d1, d2 in plages_arret:
        ws.cell(row=27+nb_plages_arret, column=12, value="du")
        ws.cell(row=27+nb_plages_arret, column=13, value=f"{d1.strftime('%d/%m')}")
        ws.cell(row=27+nb_plages_arret, column=14, value="au")
        ws.cell(row=27+nb_plages_arret, column=15, value=f"{d2.strftime('%d/%m')}")
        nb_plages_arret += 1

    # Calcul des jours fériés et du premier jour de la semaine selon le mois et l'année

    jours_feries = holidays.France(years=annee)
    nb_jours = calendar.monthrange(annee, mois)[1]
    premier_jour = date(annee, mois, 1)
    decalage = premier_jour.weekday()

    ligne = 2
    jour = 1
    colonnes = [11, 13, 15, 17, 19, 21, 23]

    # Remplissage des jours dans la cartouche du milieu 
    while jour <= nb_jours:
        for i, col in enumerate(colonnes):
            if i < decalage and jour == 1:
                continue
            if jour > nb_jours:
                break

            cell = ws.cell(row=col, column=ligne, value=jour)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            d = date(annee, mois, jour)

            if d in jours_feries:
                ws.merge_cells(start_row=col, start_column=ligne+1, end_row=col+1, end_column=ligne+3)
                cell = ws.cell(row=col, column=ligne+1, value="FERIE")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            elif d in vacances:
                if vacances[d] == (True, True):
                    ws.merge_cells(start_row=col, start_column=ligne+1, end_row=col+1, end_column=ligne+3)
                    cell = ws.cell(row=col, column=ligne+1, value=f"CP\n09:00 à 17:00")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif vacances[d] == (True, False):
                    cell = ws.cell(row=col+1, column=ligne+4, value="04:00")
                    ws.merge_cells(start_row=col, start_column=ligne+1, end_row=col+1, end_column=ligne+3)
                    cell = ws.cell(row=col, column=ligne+1, value=f"CP\n09:00 à 12:00")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif vacances[d] == (False, True):
                    cell = ws.cell(row=col, column=ligne+4, value="03:00")
                    ws.merge_cells(start_row=col, start_column=ligne+1, end_row=col+1, end_column=ligne+3)
                    cell = ws.cell(row=col, column=ligne+1, value=f"CP\n13:00 à 17:00")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif d in absences:
                if absences[d] == (True, True):
                    ws.merge_cells(start_row=col, start_column=ligne+1, end_row=col+1, end_column=ligne+3)
                    cell = ws.cell(row=col, column=ligne+1, value=f"ABS\n09:00 à 17:00")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif absences[d] == (True, False):
                    cell = ws.cell(row=col+1, column=ligne+4, value="04:00")
                    ws.merge_cells(start_row=col, start_column=ligne+1, end_row=col+1, end_column=ligne+3)
                    cell = ws.cell(row=col, column=ligne+1, value=f"ABS\n09:00 à 12:00")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif absences[d] == (False, True):
                    cell = ws.cell(row=col, column=ligne+4, value="03:00")
                    ws.merge_cells(start_row=col, start_column=ligne+1, end_row=col+1, end_column=ligne+3)
                    cell = ws.cell(row=col, column=ligne+1, value=f"ABS\n13:00 à 17:00")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif d in arret:
                if arret[d] == (True, True):
                    ws.merge_cells(start_row=col, start_column=ligne+1, end_row=col+1, end_column=ligne+3)
                    cell = ws.cell(row=col, column=ligne+1, value=f"AM\n09:00 à 17:00")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif arret[d] == (True, False):
                    cell = ws.cell(row=col+1, column=ligne+4, value="04:00")
                    ws.merge_cells(start_row=col, start_column=ligne+1, end_row=col+1, end_column=ligne+3)
                    cell = ws.cell(row=col, column=ligne+1, value=f"AM\n09:00 à 12:00")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif arret[d] == (False, True):
                    cell = ws.cell(row=col, column=ligne+4, value="03:00")
                    ws.merge_cells(start_row=col, start_column=ligne+1, end_row=col+1, end_column=ligne+3)
                    cell = ws.cell(row=col, column=ligne+1, value=f"AM\n13:00 à 17:00")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif d.weekday() in [5, 6]:
                ws.merge_cells(start_row=col, start_column=ligne+1, end_row=col+1, end_column=ligne+3)
                cell = ws.cell(row=col, column=ligne+1, value="WEEK-END")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            else:
                cell = ws.cell(row=col, column=ligne+1, value="09:00")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell = ws.cell(row=col, column=ligne+2, value="à")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell = ws.cell(row=col, column=ligne+3, value="12:00")
                cell.alignment = Alignment(horizontal="center", vertical="center")

                cell = ws.cell(row=col+1, column=ligne+1, value="13:00")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell = ws.cell(row=col+1, column=ligne+2, value="à")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell = ws.cell(row=col+1, column=ligne+3, value="17:00")
                cell.alignment = Alignment(horizontal="center", vertical="center")

                calculer_heures(ws, col, col, ligne+1, ligne+3, ligne+4)
                calculer_heures(ws, col+1, col+1, ligne+1, ligne+3, ligne+4)

            jour += 1

        decalage = 0
        ligne += 5

    # Somme des heures travaillées en semaine

    somme(ws, lignes=[11,12,13,14,15,16,17,18,19,20], col_source=6, col_resultat=2, ligne_total=25, somme="semaine")
    somme(ws, lignes=[11,12,13,14,15,16,17,18,19,20], col_source=11, col_resultat=7, ligne_total=25, somme="semaine")
    somme(ws, lignes=[11,12,13,14,15,16,17,18,19,20], col_source=16, col_resultat=12, ligne_total=25, somme="semaine")
    somme(ws, lignes=[11,12,13,14,15,16,17,18,19,20], col_source=21, col_resultat=17, ligne_total=25, somme="semaine")
    somme(ws, lignes=[11,12,13,14,15,16,17,18,19,20], col_source=26, col_resultat=22, ligne_total=25, somme="semaine")

    # Somme des heures travaillées dans le mois

    somme(ws, lignes=25, col_source=[2,7,12,17,22], col_resultat=27, ligne_total=25, somme="total")


def convertir_jours(liste):
    resultat = {}
    for j in liste:
        if j["date"] is not None:
            resultat[j["date"]] = (j["matin"], j["aprem"])
    return resultat

# Remplis un calendrier en fonction du nombre d'employés en entrée
def remplir_fiche_paie(fichier_entree, mois, annee, employes_data):
    wb = load_workbook(fichier_entree)
    modele = wb.active

    for employe in employes_data:
        ws = wb.copy_worksheet(modele)
        if employe["nom"]:
            ws.title = employe["nom"]
        else:
            ws.title = "Sans nom"
    
        vacances = convertir_jours(employe["vacances"])
        vacances_total = 0
        for jour, (mat, aprem) in vacances.items():
            if (mat and not aprem) or (not mat and aprem):
                vacances_total = vacances_total + 0.5
            elif mat and aprem:
                vacances_total = vacances_total + 1

        absences = convertir_jours(employe["absences"])
        absences_total = 0
        for jour, (mat, aprem) in absences.items():
            if (mat and not aprem) or (not mat and aprem):
                absences_total = absences_total + 0.5
            elif mat and aprem:
                absences_total = absences_total + 1

        arret = convertir_jours(employe["arret"])
        arret_total = 0
        for jour, (mat, aprem) in arret.items():
            if (mat and not aprem) or (not mat and aprem):
                arret_total = arret_total + 0.5
            elif mat and aprem:
                arret_total = arret_total + 1

        remplir_calendrier(ws, mois, annee, vacances, absences, arret, employe["nom"], employe["responsable"], employe["ddc"], employe["fdc"], vacances_total, absences_total, arret_total)

    # Sauvegarde

    wb.remove(modele)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# --- UTILISATEURS ---
USERS = st.secrets["USERS"]

# --- INITIALISATION SESSION ---
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.username = None

# --- LOGOUT ---
def logout():
    st.session_state.logged_in = False
    st.session_state.username = None
    st.rerun()

# --- LOGIN PAGE ---
def login_page():
    st.title("Connexion")

    username = st.text_input("Nom d'utilisateur")
    password = st.text_input("Mot de passe", type="password")

    if st.button("Se connecter"):
        if username in USERS and USERS[username] == password:
            st.session_state.logged_in = True
            st.session_state.username = username
            st.success("Connexion réussie")
            st.rerun()
        else:
            st.error("Identifiants incorrects")
############### Interface Streamlit #################

if not st.session_state.logged_in:
    login_page()
    st.stop()  # bloque le reste de l'app

st.button("Déconnexion", on_click=logout)

st.title("Générateur automatique de fiche de présence (Excel)")
st.write(f"Bienvenue {st.session_state.username} !")

username = st.session_state.username

if "user_data" not in st.session_state:
    st.session_state.user_data = {}

if username not in st.session_state.user_data:
    st.session_state.user_data[username] = {}

user_store = st.session_state.user_data[username]

uploaded_excel = "Fiche_Exemple.xlsx"

col1, col2 = st.columns(2)
with col1:
    user_store["mois"] = st.number_input("Mois", min_value=1, max_value=12, value=user_store.get("mois", 1), key="mois")
with col2:
    user_store["annee"] = st.number_input("Année", min_value=2000, max_value=2100, value=user_store.get("annee", 2025), key="annee")


user_store["nb_employe"] = st.number_input("Nombre d'employés :", min_value=1, max_value=30, step=1, value=user_store.get("nb_employe", 1), key="nb_employe")
employe = [f"Employé {j+1}" for j in range(user_store["nb_employe"])]

if "employes_data" not in user_store:
    user_store["employes_data"] = []

# Ajouter des employés si on augmente le nombre
while len(user_store["employes_data"]) < user_store["nb_employe"]:
    user_store["employes_data"].append({
        "nom": "",
        "responsable": "",
        "ddc": None,
        "fdc": None,
        "cdi": False,
        "vacances": [],
        "absences": [],
        "arret": []
    })

# Supprimer des employés si on diminue le nombre
while len(user_store["employes_data"]) > user_store["nb_employe"]:
    user_store["employes_data"].pop()

tabs = st.tabs(employe)

for h, tab in enumerate(tabs):
    with tab:
        emp = user_store["employes_data"][h]
        if "vacances" not in emp:
            emp["vacances"] = []
        if "absences" not in emp:
            emp["absences"] = []
        if "arret" not in emp:
            emp["arret"] = []


        st.subheader("Information Employé")
        emp["nom"] = st.text_input("NOM Prénom (Employé)", key=f"{username}_employe_nom_{h}", value=emp["nom"])
        emp["responsable"] = st.text_input("NOM Prénom (Responsable)", key=f"{username}_resp_nom_{h}", value=emp["responsable"])
        emp["ddc"] = st.date_input(f"Date de début de contrat", key=f"{username}_date_deb_contrat_{h}", format="MM/DD/YYYY", value=emp.get("ddc", None))
        emp["cdi"] = st.checkbox(f"Contrat à durée indéterminée ?", value=emp.get("cdi", False), key=f"{username}_contrat_type_{h}")
        if (emp["cdi"] == False):
            if (emp["fdc"] == "Pas de fin"):
                emp["fdc"] = st.date_input(f"Date de fin de contrat", key=f"{username}_date_fin_contrat_{h}", format="MM/DD/YYYY", value=None)
            else:
                emp["fdc"] = st.date_input(f"Date de fin de contrat", key=f"{username}_date_fin_contrat_{h}", format="MM/DD/YYYY", value=emp.get("fdc", None))
        else:
            emp["fdc"] = "Pas de fin"

        with st.expander("Congés payés"):
            st.subheader("Saisir les jours de congés payés")
            nb_jours_vac = st.number_input("Nombre de jours :", min_value=0, max_value=31, value=len(emp["vacances"]), key=f"{username}_nb_jours_vac_{h}")

            while len(emp["vacances"]) < nb_jours_vac:
                emp["vacances"].append({
                "date": None,
                "matin": False,
                "aprem": False
            })

            while len(emp["vacances"]) > nb_jours_vac:
                emp["vacances"].pop()

            for i, vac in enumerate(emp["vacances"]):
                st.markdown(f"### Jour de CP #{i+1}")
                col1, col2, col3 = st.columns(3)

                with col1:
                    vac["date"] = st.date_input(f"Date", key=f"{username}_date_cp_{h}_{i}", format="MM/DD/YYYY", value=vac["date"])
                with col2:
                    vac["matin"] = st.checkbox(f"Matin", value=vac["matin"], key=f"{username}_matin_{h}_{i}")
                with col3:
                    vac["aprem"] = st.checkbox(f"Après-midi", value=vac["aprem"], key=f"{username}_aprem_{h}_{i}")


        with st.expander("Absences"):
            st.subheader("Saisir les jours d'absences")
            nb_jours_abs = st.number_input("Nombre de jours :", min_value=0, max_value=31, value=len(emp["absences"]), key=f"{username}_nb_jours_abs_{h}")

            while len(emp["absences"]) < nb_jours_abs:
                emp["absences"].append({
                "date": None,
                "matin": False,
                "aprem": False
            })

            while len(emp["absences"]) > nb_jours_abs:
                emp["absences"].pop()

            for i, abs in enumerate(emp["absences"]):
                st.markdown(f"### Jour d'ABS #{i+1}")
                col1, col2, col3 = st.columns(3)
                with col1:
                    abs["date"] = st.date_input(f"Date", key=f"{username}_date_abs_{h}_{i}", format="MM/DD/YYYY", value=abs["date"])
                with col2:
                    abs["matin"] = st.checkbox(f"Matin", value=abs["matin"], key=f"{username}_matin_abs_{h}_{i}")
                with col3:
                    abs["aprem"] = st.checkbox(f"Après-midi", value=abs["aprem"], key=f"{username}_aprem_abs_{h}_{i}")

                
        with st.expander("Arrêts maladies"):
            st.subheader("Saisir les jours d'arrêts maladies")
            nb_jours_am = st.number_input("Nombre de jours", min_value=0, max_value=31, value=len(emp["arret"]), key=f"{username}_nb_jours_am_{h}")

            while len(emp["arret"]) < nb_jours_am:
                emp["arret"].append({
                "date": None,
                "matin": False,
                "aprem": False
            })

            while len(emp["arret"]) > nb_jours_am:
                emp["arret"].pop()

            for i, am in enumerate(emp["arret"]):
                st.markdown(f"### Jour d'AM #{i+1}")
                col1, col2, col3 = st.columns(3)
                with col1:
                    am["date"] = st.date_input(f"Date", key=f"{username}_date_am_{h}_{i}", format="MM/DD/YYYY", value=am["date"])
                with col2:
                    am["matin"] = st.checkbox(f"Matin", value=am["matin"], key=f"{username}_matin_am_{h}_{i}")
                with col3:
                    am["aprem"] = st.checkbox(f"Après-midi", value=am["aprem"], key=f"{username}_aprem_am_{h}_{i}")

# Vérifie si les cases sont bien cochées avant de générer l'excel

if st.button("Générer la fiche"): 
    erreur_type = None
    erreur_employe = None

    categories = {
        "vacances": "le congé payé",
        "absences": "l'absence",
        "arret": "l'arrêt maladie"
    }

    for idx, employe in enumerate(user_store["employes_data"], start=1):
        nom_emp = employe.get("nom", "Employé sans nom")

        for key_cat, label in categories.items():
            for jour in employe[key_cat]:
                if not jour["matin"] and not jour["aprem"]:
                    erreur_type = label
                    erreur_employe = nom_emp
                    break
            
            if erreur_type:
                break
        
        if (not employe["fdc"]):
            erreur_type = "du fin de contrat"
            erreur_employe = f"{nom_emp} (Employé {idx})"

        if (not employe["ddc"]):
            erreur_type = "du début de contrat"
            erreur_employe = f"{nom_emp} (Employé {idx})"

        if (employe["responsable"] == ""):
            erreur_type = "du responsable"
            erreur_employe = f"{nom_emp} (Employé {idx})"

        if (employe["nom"] == ""):
            erreur_type = "du nom"
            erreur_employe = f"Employé {idx}"

        if erreur_type:
            break

    if (erreur_type == "le congé payé" or erreur_type == "l'absence" or erreur_type == "l'arrêt maladie"):
        st.error(
            f"Une des deux cases 'Matin' ou 'Après-midi' pour {erreur_type} de **{erreur_employe}** n'a pas été chochée !"
        )
    elif erreur_type:
        st.error(
            f"Il manque l'information {erreur_type} pour **{erreur_employe}** !"
        )
    else:
        buffer = remplir_fiche_paie(uploaded_excel, user_store["mois"], user_store["annee"], user_store["employes_data"])

        st.success("Fiche générée avec succès !")

        st.download_button(
            "Télécharger la fiche remplie",
            data=buffer,
            file_name=f"fiche_paie_{user_store["mois"]}_{user_store["annee"]}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )